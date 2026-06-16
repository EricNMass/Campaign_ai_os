import os
import csv
from typing import List, Dict, Any
from backend.app.config import settings
from backend.app.agents.base import BaseAgent

class ReportingAgent(BaseAgent):
    def __init__(self):
        super().__init__(
            name="ReportingAgent",
            role="Data Reporter",
            system_prompt="You compile campaign metrics and logs into executive summaries."
        )

    def generate_csv_report(self, campaign_name: str, findings: List[Dict[str, Any]], output_path: str) -> str:
        """Writes CSV findings list."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Campaign Name", "Target Type", "Severity", "Description", "Remediation"])
            for finding in findings:
                writer.writerow([
                    campaign_name,
                    finding.get("target_type", ""),
                    finding.get("severity", ""),
                    finding.get("description", ""),
                    finding.get("remediation", "")
                ])
        return output_path

    def generate_excel_report(self, campaign_name: str, findings: List[Dict[str, Any]], output_path: str) -> str:
        """Writes structured Excel sheet using openpyxl."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Campaign Audit Findings"
            
            # Title block
            ws.merge_cells("A1:E1")
            ws["A1"] = f"CAMPAIGN AUDIT: {campaign_name.upper()}"
            ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            ws["A1"].fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
            ws["A1"].alignment = Alignment(horizontal="center")
            
            # Headers
            headers = ["Campaign Name", "Target Type", "Severity", "Description", "Remediation"]
            ws.append(headers)
            header_row = 2
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
                cell.alignment = Alignment(horizontal="left")
                
            # Content
            for finding in findings:
                ws.append([
                    campaign_name,
                    finding.get("target_type", ""),
                    finding.get("severity", ""),
                    finding.get("description", ""),
                    finding.get("remediation", "")
                ])
                
            # Style severity cell colors
            for r in range(3, ws.max_row + 1):
                severity_cell = ws.cell(row=r, column=3)
                sev = str(severity_cell.value).upper()
                if sev == "HIGH":
                    severity_cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
                    severity_cell.font = Font(color="991B1B", bold=True)
                elif sev == "MEDIUM":
                    severity_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    severity_cell.font = Font(color="92400E", bold=True)
                elif sev == "LOW":
                    severity_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    severity_cell.font = Font(color="065F46", bold=True)
            
            # Column widths autofit
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            wb.save(output_path)
            return output_path
        except Exception as e:
            print(f"[ReportingAgent] openpyxl failed: {e}. Defaulting to CSV naming link.")
            # Default fallback to CSV write if openpyxl fails
            return self.generate_csv_report(campaign_name, findings, output_path.replace(".xlsx", ".csv"))

    def generate_html_report(self, campaign_name: str, findings: List[Dict[str, Any]], output_path: str) -> str:
        """Generates simple, responsive campaign findings HTML report page."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        rows_html = ""
        for index, finding in enumerate(findings):
            sev = finding.get("severity", "").upper()
            color_class = "bg-red-500/10 text-red-400 border-red-500/20" if sev == "HIGH" else (
                "bg-amber-500/10 text-amber-400 border-amber-500/20" if sev == "MEDIUM" else
                "bg-emerald-500/10 text-emerald-400 border-emerald-500/20"
            )
            rows_html += f"""
            <tr class="border-b border-white/5 hover:bg-white/2 transition-colors">
                <td class="px-6 py-4 font-semibold text-white">{finding.get('target_type', '')}</td>
                <td class="px-6 py-4">
                    <span class="px-2.5 py-1 text-[10px] font-black uppercase tracking-wider rounded-full border {color_class}">
                        {sev}
                    </span>
                </td>
                <td class="px-6 py-4 text-slate-300">{finding.get('description', '')}</td>
                <td class="px-6 py-4 text-indigo-400 italic">{finding.get('remediation', '')}</td>
            </tr>
            """

        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{campaign_name} Audit Report</title>
    <script src="https://cdn.tailwindcss.com"></script>
</head>
<body class="bg-slate-950 text-slate-100 font-sans p-8">
    <div class="max-w-6xl mx-auto">
        <header class="mb-10 pb-6 border-b border-white/10 flex justify-between items-center">
            <div>
                <h1 class="text-3xl font-black text-white tracking-tight">{campaign_name.upper()}</h1>
                <p class="text-xs text-indigo-400 font-bold uppercase tracking-widest mt-1">Campaign Integrity Audit Summary</p>
            </div>
            <div class="text-right text-slate-400 text-xs">
                <p>Run Time: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
                <p>Status: <span class="text-emerald-400 font-bold">Audit Complete</span></p>
            </div>
        </header>
        
        <div class="overflow-x-auto bg-white/2 border border-white/10 rounded-2xl">
            <table class="w-full text-left text-sm">
                <thead>
                    <tr class="bg-white/5 border-b border-white/10 text-slate-400 uppercase text-xs tracking-wider">
                        <th class="px-6 py-4 font-bold">Component</th>
                        <th class="px-6 py-4 font-bold">Severity</th>
                        <th class="px-6 py-4 font-bold">Description</th>
                        <th class="px-6 py-4 font-bold">Remediation Action</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html if findings else '<tr><td colspan="4" class="px-6 py-12 text-center text-slate-400 italic">No defects identified. Campaign parameters fully compliant.</td></tr>'}
                </tbody>
            </table>
        </div>
    </div>
</body>
</html>
"""
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        return output_path

    def generate_pdf_report(self, campaign_name: str, findings: List[Dict[str, Any]], output_path: str) -> str:
        """Generates executive PDF report using reportlab flowables."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            doc = SimpleDocTemplate(output_path, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            # Setup styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Heading1'],
                fontSize=20,
                leading=24,
                textColor=colors.HexColor('#1E1B4B'),
                spaceAfter=15
            )
            sub_style = ParagraphStyle(
                'SubStyle',
                parent=styles['Normal'],
                fontSize=9,
                leading=12,
                textColor=colors.HexColor('#4F46E5'),
                spaceAfter=25
            )
            body_style = ParagraphStyle(
                'BodyStyle',
                parent=styles['Normal'],
                fontSize=10,
                leading=14,
                textColor=colors.HexColor('#334155')
            )
            
            story.append(Paragraph(f"CAMPAIGN INTEGRITY REPORT: {campaign_name.upper()}", title_style))
            story.append(Paragraph(f"Autonomous Audit Completed at: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", sub_style))
            story.append(Spacer(1, 15))
            
            # Format Table
            data = [["Component", "Severity", "Issue Description", "Remediation"]]
            for f in findings:
                data.append([
                    Paragraph(f.get("target_type", ""), body_style),
                    Paragraph(f.get("severity", ""), body_style),
                    Paragraph(f.get("description", ""), body_style),
                    Paragraph(f.get("remediation", ""), body_style)
                ])
                
            t = Table(data, colWidths=[80, 60, 210, 180])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 8),
                ('TOPPADDING', (0,0), (-1,0), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,1), (-1,-1), 6),
                ('TOPPADDING', (0,1), (-1,-1), 6),
            ]))
            story.append(t)
            
            doc.build(story)
            return output_path
        except Exception as e:
            print(f"[ReportingAgent] reportlab failed: {e}. Fallback: Generating HTML copy.")
            return self.generate_html_report(campaign_name, findings, output_path.replace(".pdf", ".html"))

from datetime import datetime
